"""
playback.master_data_reporter
-----------------------------
Append-only master workbook for replay runtime data.
"""
from __future__ import annotations

import json
import os
import re
import threading
import time
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import PlaybackConfig, PlaybackResult, StepResult, WorkflowStep


MASTER_WORKBOOK_NAME = "Replay_Master_Data.xlsx"

RUN_HISTORY_SHEET = "Run History"
FAILED_RUNS_SHEET = "Failed Runs Only"
PLAN_SUMMARY_SHEET = "Plan Selection Summary"

_WRITE_LOCK = threading.RLock()

_FILL_HEADER = PatternFill("solid", fgColor="1F4E78")
_FILL_SUCCESS = PatternFill("solid", fgColor="00B050")
_FILL_FAILED = PatternFill("solid", fgColor="C00000")
_FILL_WARNING = PatternFill("solid", fgColor="FFC000")
_FILL_DEFAULT = PatternFill("solid", fgColor="FFFFFF")

_FONT_HEADER = Font(bold=True, color="FFFFFF")
_FONT_STATUS_LIGHT = Font(bold=True, color="FFFFFF")
_FONT_STATUS_DARK = Font(bold=True, color="000000")
_FONT_NORMAL = Font(color="000000")

_THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)

_METADATA_HEADERS = [
    "ID",
    "Workflow Name",
    "JSON File Name",
    "Run Date",
    "Run Time",
    "Total Duration",
    "Replay Status",
    "Failed Step Count",
    "Failed Step Names",
    "Failure Reason",
    "Screenshots Created Count",
    "Screenshot File Names",
    "Excel Report Name",
    "Faker Seed Used",
    "Browser Used",
    "Environment URL",
]

_STANDARD_DATA_HEADERS = [
    "First Name",
    "Middle Initial",
    "Last Name",
    "DOB",
    "Gender",
    "Marital Status",
    "Prefix",
    "Suffix",
    "Email",
    "Address1",
    "Address2",
    "City",
    "County",
    "State",
    "Zip Code",
    "Work Phone",
    "Sponsor SSN",
    "Employee ID",
    "Date of Hire",
    "Effective Date",
    "Retirement Date",
    "Job Title",
    "Bargaining Unit",
    "Tobacco Use",
    "Medical Plan Selected",
    "Dental Plan Selected",
    "Waived Plans",
    "Payment Option",
    "Enrollment Period",
    "Billing Location",
    "Employee Class",
    "Plan Toggles Selected",
]

_AUDIT_HEADERS = [
    "Runtime Field Count",
    "All Runtime Values",
]

_PLAN_SUMMARY_HEADERS = [
    "ID",
    "Workflow Name",
    "Run Date",
    "Medical Plan Selected",
    "Dental Plan Selected",
    "Waived Plans",
    "Payment Option",
    "Enrollment Period",
    "Billing Location",
    "Employee Class",
    "Plan Toggles Selected",
]


def build_master_data_path(reports_dir: str = "reports") -> str:
    """Return the persistent master workbook path."""
    return os.path.join(reports_dir, MASTER_WORKBOOK_NAME)


def append_master_data_run(
    cfg: PlaybackConfig,
    result: PlaybackResult,
    pairs: Sequence[Tuple[WorkflowStep, StepResult]],
    runtime_fields: Sequence[Dict[str, Any]],
    reports_dir: str = "reports",
    retries: int = 8,
    retry_delay_seconds: float = 0.35,
) -> str:
    """Append one replay run to the master workbook and return the saved path."""
    workbook_path = build_master_data_path(reports_dir=reports_dir)
    os.makedirs(os.path.dirname(os.path.abspath(workbook_path)), exist_ok=True)

    with _WRITE_LOCK:
        last_error: Optional[BaseException] = None
        for attempt_index in range(max(1, retries)):
            try:
                with _WorkbookAppendLock(workbook_path):
                    return _append_once(workbook_path, cfg, result, pairs, runtime_fields)
            except (PermissionError, OSError) as exc:
                last_error = exc
                if not _is_retryable_file_error(exc) or attempt_index >= retries - 1:
                    raise
                time.sleep(retry_delay_seconds * (attempt_index + 1))

        if last_error:
            raise last_error
        raise RuntimeError("Master workbook append failed without an exception")


def update_master_excel_report_name(
    workbook_path: str,
    run_id: str,
    excel_report_path: str,
    retries: int = 8,
    retry_delay_seconds: float = 0.35,
) -> bool:
    """Update Excel Report Name for a previously appended master row."""
    if not workbook_path or not run_id or not excel_report_path:
        return False

    with _WRITE_LOCK:
        for attempt_index in range(max(1, retries)):
            try:
                with _WorkbookAppendLock(workbook_path):
                    return _update_excel_report_name_once(workbook_path, run_id, excel_report_path)
            except (PermissionError, OSError) as exc:
                if not _is_retryable_file_error(exc) or attempt_index >= retries - 1:
                    raise
                time.sleep(retry_delay_seconds * (attempt_index + 1))
    return False


class _WorkbookAppendLock:
    """Small cross-thread/process lock to prevent duplicate ID allocation."""

    def __init__(self, workbook_path: str, timeout_seconds: float = 15.0) -> None:
        self._lock_path = f"{os.path.abspath(workbook_path)}.lock"
        self._timeout_seconds = max(1.0, timeout_seconds)
        self._handle: Optional[int] = None

    def __enter__(self) -> "_WorkbookAppendLock":
        deadline = time.time() + self._timeout_seconds
        while time.time() < deadline:
            try:
                self._handle = os.open(self._lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
                payload = f"pid={os.getpid()} created_at={datetime.now().isoformat()}\n"
                os.write(self._handle, payload.encode("utf-8", errors="replace"))
                return self
            except FileExistsError:
                self._remove_stale_lock()
                time.sleep(0.2)
        raise TimeoutError(f"Could not acquire master workbook lock: {self._lock_path}")

    def __exit__(self, exc_type: Any, exc: Any, traceback: Any) -> None:
        if self._handle is not None:
            try:
                os.close(self._handle)
            except OSError:
                pass
            self._handle = None
        try:
            os.remove(self._lock_path)
        except FileNotFoundError:
            pass
        except OSError:
            pass

    def _remove_stale_lock(self) -> None:
        try:
            age_seconds = time.time() - os.path.getmtime(self._lock_path)
        except OSError:
            return
        if age_seconds < 120.0:
            return
        try:
            os.remove(self._lock_path)
        except OSError:
            pass


def _append_once(
    workbook_path: str,
    cfg: PlaybackConfig,
    result: PlaybackResult,
    pairs: Sequence[Tuple[WorkflowStep, StepResult]],
    runtime_fields: Sequence[Dict[str, Any]],
) -> str:
    workbook = _load_or_create_workbook(workbook_path)
    run_sheet = _ensure_sheet(workbook, RUN_HISTORY_SHEET)
    failed_sheet = _ensure_sheet(workbook, FAILED_RUNS_SHEET)
    plan_sheet = _ensure_sheet(workbook, PLAN_SUMMARY_SHEET)

    row_payload = _build_row_payload(cfg, result, pairs, runtime_fields)
    dynamic_headers = _dynamic_field_headers(row_payload)
    run_headers = _METADATA_HEADERS + _STANDARD_DATA_HEADERS + dynamic_headers + _AUDIT_HEADERS

    _ensure_headers(run_sheet, run_headers)
    _ensure_headers(failed_sheet, run_headers)
    _ensure_headers(plan_sheet, _PLAN_SUMMARY_HEADERS)

    next_id = _next_run_id(run_sheet)
    row_payload["ID"] = next_id
    result.master_data_run_id = next_id

    _append_row(run_sheet, row_payload)
    if row_payload.get("Replay Status") == "Failed":
        _append_row(failed_sheet, row_payload)

    plan_payload = {header: row_payload.get(header, "") for header in _PLAN_SUMMARY_HEADERS}
    _append_row(plan_sheet, plan_payload)

    for worksheet in (run_sheet, failed_sheet, plan_sheet):
        _format_sheet(worksheet)

    workbook.save(workbook_path)
    return os.path.abspath(workbook_path)


def _update_excel_report_name_once(workbook_path: str, run_id: str, excel_report_path: str) -> bool:
    if not os.path.exists(workbook_path):
        return False

    workbook = openpyxl.load_workbook(workbook_path)
    updated = False
    report_name = os.path.basename(str(excel_report_path or ""))
    for sheet_name in (RUN_HISTORY_SHEET, FAILED_RUNS_SHEET):
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        id_column = _header_column(worksheet, "ID")
        report_column = _header_column(worksheet, "Excel Report Name")
        if not id_column or not report_column:
            continue
        for row_index in range(2, worksheet.max_row + 1):
            if str(worksheet.cell(row=row_index, column=id_column).value or "").strip() != str(run_id).strip():
                continue
            cell = worksheet.cell(row=row_index, column=report_column, value=report_name)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = _THIN_BORDER
            updated = True
            break
        _format_sheet(worksheet)

    if updated:
        workbook.save(workbook_path)
    return updated


def _load_or_create_workbook(workbook_path: str) -> openpyxl.Workbook:
    if os.path.exists(workbook_path):
        return openpyxl.load_workbook(workbook_path)

    workbook = openpyxl.Workbook()
    workbook.active.title = RUN_HISTORY_SHEET
    return workbook


def _ensure_sheet(workbook: openpyxl.Workbook, title: str) -> Any:
    if title in workbook.sheetnames:
        return workbook[title]
    return workbook.create_sheet(title)


def _ensure_headers(worksheet: Any, required_headers: Sequence[str]) -> List[str]:
    existing_headers = _read_headers(worksheet)
    if not existing_headers:
        for column_index, header in enumerate(required_headers, start=1):
            worksheet.cell(row=1, column=column_index, value=header)
        return list(required_headers)

    for header in required_headers:
        if header in existing_headers:
            continue
        existing_headers.append(header)
        worksheet.cell(row=1, column=len(existing_headers), value=header)
    return existing_headers


def _read_headers(worksheet: Any) -> List[str]:
    if worksheet.max_row < 1:
        return []
    headers: List[str] = []
    for column_index in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=1, column=column_index).value
        text = str(value or "").strip()
        if text:
            headers.append(text)
    return headers


def _next_run_id(worksheet: Any) -> str:
    max_id = 0
    for row_index in range(2, worksheet.max_row + 1):
        raw_value = worksheet.cell(row=row_index, column=1).value
        digits = re.sub(r"\D+", "", str(raw_value or ""))
        if not digits:
            continue
        try:
            max_id = max(max_id, int(digits))
        except ValueError:
            continue
    return f"{max_id + 1:03d}"


def _append_row(worksheet: Any, row_payload: Dict[str, Any]) -> int:
    headers = _read_headers(worksheet)
    row_index = worksheet.max_row + 1 if worksheet.max_row >= 1 else 2
    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=row_index, column=column_index, value=_cell_value(row_payload.get(header, "")))
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = _THIN_BORDER
        cell.font = _FONT_NORMAL
        if header == "ID":
            cell.number_format = "@"
        if header == "Replay Status":
            _style_status_cell(cell)
    return row_index


def _format_sheet(worksheet: Any) -> None:
    if worksheet.max_row < 1:
        return

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_index in range(1, worksheet.max_column + 1):
        header_cell = worksheet.cell(row=1, column=column_index)
        header_cell.font = _FONT_HEADER
        header_cell.fill = _FILL_HEADER
        header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_cell.border = _THIN_BORDER

    for row_index in range(2, worksheet.max_row + 1):
        status_column = _header_column(worksheet, "Replay Status")
        if status_column:
            _style_status_cell(worksheet.cell(row=row_index, column=status_column))

    for column_index in range(1, worksheet.max_column + 1):
        letter = get_column_letter(column_index)
        worksheet.column_dimensions[letter].width = _column_width(worksheet, column_index)


def _header_column(worksheet: Any, header: str) -> int:
    for column_index, value in enumerate(_read_headers(worksheet), start=1):
        if value == header:
            return column_index
    return 0


def _style_status_cell(cell: Any) -> None:
    status = str(cell.value or "").strip().lower()
    if status == "success":
        cell.fill = _FILL_SUCCESS
        cell.font = _FONT_STATUS_LIGHT
    elif status == "failed":
        cell.fill = _FILL_FAILED
        cell.font = _FONT_STATUS_LIGHT
    elif status in {"warning", "partial", "passed with warnings"}:
        cell.fill = _FILL_WARNING
        cell.font = _FONT_STATUS_DARK
    else:
        cell.fill = _FILL_DEFAULT
        cell.font = _FONT_NORMAL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _THIN_BORDER


def _column_width(worksheet: Any, column_index: int) -> float:
    max_length = 10
    for row_index in range(1, worksheet.max_row + 1):
        value = worksheet.cell(row=row_index, column=column_index).value
        if value is None:
            continue
        text = str(value)
        if len(text) > 100:
            text = text[:100]
        max_length = max(max_length, len(text) + 2)
    return float(min(max(max_length, 10), 60))


def _build_row_payload(
    cfg: PlaybackConfig,
    result: PlaybackResult,
    pairs: Sequence[Tuple[WorkflowStep, StepResult]],
    runtime_fields: Sequence[Dict[str, Any]],
) -> Dict[str, Any]:
    completed_at = datetime.now()
    payload = result.to_dict()
    field_values, dynamic_values, runtime_summary = _extract_runtime_values(runtime_fields)
    screenshot_names = _screenshot_names(result, pairs)
    failed_names = _failed_step_names(pairs)
    workflow_name, json_file_name = _workflow_identity(cfg)

    row_payload: Dict[str, Any] = {
        "Workflow Name": workflow_name,
        "JSON File Name": json_file_name,
        "Run Date": completed_at.strftime("%Y-%m-%d"),
        "Run Time": completed_at.strftime("%H:%M:%S"),
        "Total Duration": f"{float(result.duration_seconds or 0.0):.2f}s",
        "Replay Status": _replay_status(payload),
        "Failed Step Count": int(result.steps_failed or 0),
        "Failed Step Names": ", ".join(failed_names),
        "Failure Reason": str(result.error or ""),
        "Screenshots Created Count": len(screenshot_names),
        "Screenshot File Names": ", ".join(screenshot_names),
        "Excel Report Name": os.path.basename(str(result.excel_report_path or "")),
        "Faker Seed Used": _data_generation_value(result, "seed"),
        "Browser Used": str(cfg.execution_profile.get("browser") or cfg.execution_profile.get("browser_name") or "chromium"),
        "Environment URL": str(cfg.start_url or ""),
        "Runtime Field Count": len(runtime_fields),
        "All Runtime Values": runtime_summary,
    }

    for header in _STANDARD_DATA_HEADERS:
        row_payload[header] = field_values.get(header, "")
    row_payload.update(dynamic_values)
    return row_payload


def _workflow_identity(cfg: PlaybackConfig) -> Tuple[str, str]:
    profile = cfg.execution_profile if isinstance(cfg.execution_profile, dict) else {}
    workflow_data = cfg.workflow_data if isinstance(cfg.workflow_data, dict) else {}
    raw_name = str(
        profile.get("_workflow_name")
        or profile.get("workflow_name")
        or workflow_data.get("file_name")
        or workflow_data.get("name")
        or ""
    ).strip()
    json_file_name = os.path.basename(raw_name) if raw_name else ""
    if json_file_name and not json_file_name.lower().endswith(".json"):
        json_file_name = f"{json_file_name}.json"
    workflow_name = os.path.splitext(json_file_name)[0] if json_file_name else str(workflow_data.get("name") or "")
    return workflow_name, json_file_name


def _replay_status(payload: Dict[str, Any]) -> str:
    outcome = str(payload.get("qa_outcome") or "").strip().lower()
    raw_status = str(payload.get("status") or "").strip().lower()
    failed_steps = int(payload.get("steps_failed", 0) or 0)
    error = str(payload.get("error") or "").strip()

    if outcome == "failed" or raw_status != "finished" or failed_steps > 0 or error:
        return "Failed"
    if outcome == "passed with warnings":
        return "Warning"
    return "Success"


def _data_generation_value(result: PlaybackResult, key: str) -> Any:
    data = result.data_generation if isinstance(result.data_generation, dict) else {}
    return data.get(key, "")


def _failed_step_names(pairs: Sequence[Tuple[WorkflowStep, StepResult]]) -> List[str]:
    names: List[str] = []
    for step, step_result in pairs:
        if step_result.success or step_result.skipped:
            continue
        label = step.display_label or step_result.step_label
        if label and label not in names:
            names.append(label)
    return names


def _screenshot_names(result: PlaybackResult, pairs: Sequence[Tuple[WorkflowStep, StepResult]]) -> List[str]:
    paths: List[str] = []
    if result.screenshot_path:
        paths.append(str(result.screenshot_path))
    for _step, step_result in pairs:
        if step_result.screenshot_path:
            paths.append(str(step_result.screenshot_path))
    for discrepancy in result.discrepancies or []:
        if discrepancy.screenshot_path:
            paths.append(str(discrepancy.screenshot_path))

    names: List[str] = []
    for path in paths:
        name = os.path.basename(path)
        if name and name not in names:
            names.append(name)
    return names


def _extract_runtime_values(runtime_fields: Sequence[Dict[str, Any]]) -> Tuple[Dict[str, str], Dict[str, str], str]:
    standard_values: Dict[str, str] = {header: "" for header in _STANDARD_DATA_HEADERS}
    dynamic_values: Dict[str, str] = {}
    summary_parts: List[str] = []
    plan_toggles: List[str] = []
    waived_plans: List[str] = []

    for record in runtime_fields:
        value = _clean_runtime_value(record.get("value", ""))
        if not value:
            continue

        label = _display_label(record)
        status = str(record.get("status") or "").strip()
        if label:
            summary_parts.append(f"{label}={value}" + (f" ({status})" if status else ""))

        category = str(record.get("category") or "").strip().lower()
        plan_context = str(record.get("plan_context") or "").strip()
        plan_context_norm = _compact(plan_context)

        if category == "plan_selection":
            plan_value = _merge_plan_context(plan_context, value)
            if "medical" in plan_context_norm:
                _merge_header_value(standard_values, "Medical Plan Selected", plan_value)
            elif "dental" in plan_context_norm:
                _merge_header_value(standard_values, "Dental Plan Selected", plan_value)
            else:
                _append_unique(plan_toggles, plan_value)
            continue

        if category == "plan_decline":
            toggle_value = _merge_plan_context(plan_context, value)
            _append_unique(plan_toggles, toggle_value)
            if "declined" in value.lower() and "not declined" not in value.lower():
                _append_unique(waived_plans, plan_context or label or value)
            continue

        if category == "payment":
            _merge_header_value(standard_values, "Payment Option", value)
            continue

        header = _canonical_header(record)
        if header:
            _merge_header_value(standard_values, header, value)
            continue

        dynamic_header = _dynamic_header(label)
        if dynamic_header:
            _merge_header_value(dynamic_values, dynamic_header, value)

    if waived_plans:
        standard_values["Waived Plans"] = ", ".join(waived_plans)
    if plan_toggles:
        standard_values["Plan Toggles Selected"] = ", ".join(plan_toggles)

    runtime_summary = _truncate("; ".join(summary_parts), 32700)
    return standard_values, dynamic_values, runtime_summary


def _canonical_header(record: Dict[str, Any]) -> str:
    blob = _compact(" ".join(str(record.get(key) or "") for key in ("label", "name", "id", "text", "input_type")))

    if any(token in blob for token in ("firstname", "givenname", "fname")):
        return "First Name"
    if any(token in blob for token in ("middleinitial", "middleinit", "middlename", "mname")):
        return "Middle Initial"
    if any(token in blob for token in ("lastname", "surname", "familyname", "lname")):
        return "Last Name"
    if any(token in blob for token in ("dateofbirth", "birthdate", "dob")):
        return "DOB"
    if "gender" in blob or blob.endswith("sex"):
        return "Gender"
    if "maritalstatus" in blob:
        return "Marital Status"
    if "prefix" in blob or "salutation" in blob:
        return "Prefix"
    if "suffix" in blob:
        return "Suffix"
    if "email" in blob:
        return "Email"
    if any(token in blob for token in ("address1", "addressline1", "streetaddress", "streetaddr")):
        return "Address1"
    if any(token in blob for token in ("address2", "addressline2", "apt", "suite")):
        return "Address2"
    if "city" in blob:
        return "City"
    if "county" in blob:
        return "County"
    if "state" in blob or "province" in blob:
        return "State"
    if any(token in blob for token in ("zipcode", "postalcode", "zip")) and "extension" not in blob and "zipext" not in blob:
        return "Zip Code"
    if any(token in blob for token in ("workphone", "phone", "telephone", "altphone")) and "fax" not in blob:
        return "Work Phone"
    if "ssn" in blob or "socialsecurity" in blob:
        return "Sponsor SSN"
    if any(token in blob for token in ("employeeid", "empid", "staffid", "workerid")):
        return "Employee ID"
    if any(token in blob for token in ("dateofhire", "hiredate", "employmentdate", "startdate")):
        return "Date of Hire"
    if any(token in blob for token in ("effectivedate", "coverageeffective")):
        return "Effective Date"
    if any(token in blob for token in ("retirementdate", "retiredate", "terminationdate", "termdate", "enddate")):
        return "Retirement Date"
    if any(token in blob for token in ("jobtitle", "positiontitle", "occupation")):
        return "Job Title"
    if any(token in blob for token in ("bargainingunit", "unioncode")):
        return "Bargaining Unit"
    if "tobacco" in blob or "tobaco" in blob or "smoker" in blob:
        return "Tobacco Use"
    if any(token in blob for token in ("billinglocation", "subgroup")):
        return "Billing Location"
    if any(token in blob for token in ("employeeclass", "classid", "dropdowclasses", "dropdownclasses")):
        return "Employee Class"
    if any(token in blob for token in ("enrollmentperiod", "enrollmentwindow", "planyear", "dropdownplanyear")):
        return "Enrollment Period"
    if "payment" in blob or blob in {"rdno", "rdyes"}:
        return "Payment Option"
    return ""


def _dynamic_field_headers(row_payload: Dict[str, Any]) -> List[str]:
    reserved = set(_METADATA_HEADERS + _STANDARD_DATA_HEADERS + _AUDIT_HEADERS)
    headers = [header for header in row_payload.keys() if header not in reserved and header != "ID"]
    return sorted(headers)


def _display_label(record: Dict[str, Any]) -> str:
    for key in ("label", "text", "name", "id"):
        value = str(record.get(key) or "").strip()
        if value:
            return re.sub(r"\s+", " ", value).strip(" :")[:90]
    return ""


def _dynamic_header(label: str) -> str:
    cleaned = re.sub(r"\s+", " ", str(label or "")).strip(" :")
    if not cleaned:
        return ""
    if cleaned in set(_METADATA_HEADERS + _STANDARD_DATA_HEADERS + _AUDIT_HEADERS):
        return ""
    return cleaned[:90]


def _merge_header_value(target: Dict[str, str], header: str, value: str) -> None:
    if not header or not value:
        return
    existing = str(target.get(header) or "").strip()
    if not existing:
        target[header] = value
        return
    parts = [part.strip() for part in existing.split(",") if part.strip()]
    if value not in parts and value != existing:
        target[header] = f"{existing}, {value}"


def _append_unique(target: List[str], value: str) -> None:
    cleaned = str(value or "").strip()
    if cleaned and cleaned not in target:
        target.append(cleaned)


def _merge_plan_context(plan_context: str, value: str) -> str:
    if not plan_context:
        return value
    if plan_context.lower() in value.lower():
        return value
    return f"{plan_context}: {value}"


def _clean_runtime_value(value: Any) -> str:
    text = str(value or "").strip()
    for prefix in ("recorded:", "selected:"):
        if text.lower().startswith(prefix):
            text = text[len(prefix):].strip()
    return _truncate(re.sub(r"\s+", " ", text), 1200)


def _compact(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        value = json.dumps(value, ensure_ascii=True, default=str)
    if isinstance(value, str):
        return _truncate(value, 32700)
    return value


def _truncate(value: str, limit: int) -> str:
    text = str(value or "")
    if len(text) <= limit:
        return text
    return text[: max(0, limit - 3)] + "..."


def _is_retryable_file_error(exc: BaseException) -> bool:
    if isinstance(exc, PermissionError):
        return True
    message = str(exc).lower()
    return any(token in message for token in ("permission denied", "being used", "locked", "access is denied"))
