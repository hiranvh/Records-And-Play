"""
playback.excel_reporter
-----------------------
Writes a colour-coded Excel workbook summarising one playback run.

Each row represents one workflow step.  Columns:
  Step #  |  Field Label  |  Type  |  Faker Value  |  Status  |  Reason  |  Screenshot  |  Timestamp

Status colour coding:
  PASS  → green
  FAIL  → red
  SKIP  → yellow

The worksheet also includes a two-row header block with workflow name,
run timestamp, and total duration so reports are self-contained.
"""
from __future__ import annotations

import json
import os
import re
from datetime import datetime
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ── Colours ───────────────────────────────────────────────────────────────────

_FILL_HEADER  = PatternFill("solid", fgColor="2F4F8F")   # dark blue
_FILL_META    = PatternFill("solid", fgColor="D9E1F2")   # light blue
_FILL_PASS    = PatternFill("solid", fgColor="C6EFCE")   # green
_FILL_FAIL    = PatternFill("solid", fgColor="FFC7CE")   # red
_FILL_SKIP    = PatternFill("solid", fgColor="FFEB9C")   # yellow

_FONT_HEADER  = Font(bold=True, color="FFFFFF", size=11)
_FONT_META_LBL= Font(bold=True, size=10)
_FONT_NORMAL  = Font(size=10)

_THIN_BORDER  = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_COL_HEADERS = [
    "Step #",
    "Field Label",
    "Step Type",
    "Faker Value Used",
    "Status",
    "Reason / Message",
    "Screenshot Path",
    "Timestamp",
]

_COL_WIDTHS = [8, 32, 14, 36, 10, 55, 55, 22]

_SEVERITY_FILLS = {
    "blocker": PatternFill("solid", fgColor="FFC7CE"),
    "warning": PatternFill("solid", fgColor="FFEB9C"),
    "info": PatternFill("solid", fgColor="D9E1F2"),
}


class ExcelReporter:
    """
    Writes a structured Excel report for one playback run.

    Usage::

        reporter = ExcelReporter("reports/run_001.xlsx")
        reporter.write(
            rows           = [...],          # list[dict] — one per step
            workflow_name  = "Enrollment",
            duration_seconds = 42.3,
        )

    Each dict in ``rows`` must have these keys:
        step_index   int
        label        str   — human-readable step label
        step_type    str   — input / select / click / …
        faker_value  str   — value used (empty for non-input steps)
        status       str   — PASS | FAIL | SKIP
        reason       str   — failure message or blank
        screenshot   str   — file path or blank
        timestamp    str   — ISO-format datetime string
    """

    def __init__(self, path: str) -> None:
        self.path = path

    # ── Public ────────────────────────────────────────────────────────────────

    def write(
        self,
        rows: List[Dict[str, Any]],
        workflow_name: str = "",
        duration_seconds: float = 0.0,
    ) -> str:
        """Write the workbook and return the saved absolute path."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Playback Results"

        self._write_meta(ws, workflow_name, duration_seconds, rows)
        self._write_header(ws)
        self._write_rows(ws, rows)
        self._apply_column_widths(ws)
        ws.freeze_panes = "A5"   # freeze meta + header rows

        os.makedirs(os.path.dirname(os.path.abspath(self.path)), exist_ok=True)
        wb.save(self.path)
        return os.path.abspath(self.path)

    # ── Private ───────────────────────────────────────────────────────────────

    def _write_meta(
        self,
        ws: Any,
        workflow_name: str,
        duration: float,
        rows: List[Dict[str, Any]],
    ) -> None:
        """Rows 1-3: summary block."""
        total  = len(rows)
        passed = sum(1 for r in rows if str(r.get("status", "")).upper() == "PASS")
        failed = sum(1 for r in rows if str(r.get("status", "")).upper() == "FAIL")
        skipped= sum(1 for r in rows if str(r.get("status", "")).upper() == "SKIP")

        meta_pairs = [
            ("Workflow",   workflow_name),
            ("Run At",     datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Duration",   f"{duration:.1f}s"),
            ("Total",      str(total)),
            ("Passed",     str(passed)),
            ("Failed",     str(failed)),
            ("Skipped",    str(skipped)),
        ]

        # Row 1: labels, Row 2: values, alternating columns
        labels = [p[0] for p in meta_pairs]
        values = [p[1] for p in meta_pairs]

        # Write labels in row 1
        for col, lbl in enumerate(labels, start=1):
            cell = ws.cell(row=1, column=col, value=lbl)
            cell.font = _FONT_META_LBL
            cell.fill = _FILL_META
            cell.alignment = Alignment(horizontal="center")

        # Write values in row 2
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=2, column=col, value=val)
            cell.font = _FONT_NORMAL
            cell.fill = _FILL_META
            cell.alignment = Alignment(horizontal="center")

        ws.append([])  # Row 3: blank spacer before data

    def _write_header(self, ws: Any) -> None:
        """Row 4: column headers."""
        ws.append(_COL_HEADERS)
        row_idx = ws.max_row
        ws.row_dimensions[row_idx].height = 22
        for col_idx, _ in enumerate(_COL_HEADERS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font  = _FONT_HEADER
            cell.fill  = _FILL_HEADER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _THIN_BORDER

    def _write_rows(self, ws: Any, rows: List[Dict[str, Any]]) -> None:
        """Rows 5+: one row per step."""
        for r in rows:
            status = str(r.get("status", "")).upper()
            fill = (
                _FILL_PASS  if status == "PASS" else
                _FILL_FAIL  if status == "FAIL" else
                _FILL_SKIP
            )
            row_data = [
                r.get("step_index", ""),
                r.get("label", ""),
                r.get("step_type", ""),
                r.get("faker_value", ""),
                status,
                r.get("reason", ""),
                r.get("screenshot", ""),
                r.get("timestamp", ""),
            ]
            ws.append(row_data)
            row_num = ws.max_row
            for col_idx in range(1, len(_COL_HEADERS) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = _THIN_BORDER
                cell.font = _FONT_NORMAL

    def _apply_column_widths(self, ws: Any) -> None:
        for col_idx, width in enumerate(_COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_report_path(workflow_name: str, reports_dir: str = "reports") -> str:
    """
    Build a timestamped Excel report path.

    Example::
        build_report_path("Enrollment")
        # → "reports/Enrollment_20260423_143012.xlsx"
    """
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = re.sub(r"[^\w\-]+", "_", workflow_name or "run")[:40]
    return os.path.join(reports_dir, f"{safe_name}_{ts}.xlsx")


def build_discrepancy_report_paths(
    workflow_name: str,
    reports_dir: str = "reports",
) -> Dict[str, str]:
    """Build timestamped output paths for QA discrepancy JSON + Excel reports."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = re.sub(r"[^\w\-]+", "_", workflow_name or "run")[:40]
    base = os.path.join(reports_dir, f"{safe_name}_qa_discrepancy_{ts}")
    return {
        "excel": f"{base}.xlsx",
        "json": f"{base}.json",
    }


def export_discrepancy_reports(
    result: Dict[str, Any],
    workflow_name: str = "",
    reports_dir: str = "reports",
) -> Dict[str, Any]:
    """Write discrepancy-focused JSON + Excel artifacts and return saved paths."""
    if not isinstance(result, dict):
        return {}

    paths = build_discrepancy_report_paths(workflow_name=workflow_name, reports_dir=reports_dir)
    os.makedirs(os.path.dirname(os.path.abspath(paths["excel"])), exist_ok=True)

    discrepancies_raw = result.get("discrepancies") or []
    if not isinstance(discrepancies_raw, list):
        discrepancies_raw = []

    discrepancies: List[Dict[str, Any]] = []
    for raw in discrepancies_raw:
        if not isinstance(raw, dict):
            continue
        kind = str(raw.get("kind") or "")
        severity = str(raw.get("severity") or "").strip().lower()
        if severity not in {"blocker", "warning", "info"}:
            severity = _severity_from_kind(kind)
        discrepancies.append({
            "severity": severity,
            "kind": kind,
            "field_label": str(raw.get("field_label") or ""),
            "field_id": str(raw.get("field_id") or ""),
            "field_name": str(raw.get("field_name") or ""),
            "tag": str(raw.get("tag") or ""),
            "input_type": str(raw.get("input_type") or ""),
            "field_required": bool(raw.get("field_required", False)),
            "field_visible": bool(raw.get("field_visible", True)),
            "field_actionable": bool(raw.get("field_actionable", True)),
            "expected_field_id": str(raw.get("expected_field_id") or ""),
            "expected_field_name": str(raw.get("expected_field_name") or ""),
            "expected_field_label": str(raw.get("expected_field_label") or ""),
            "live_field_id": str(raw.get("live_field_id") or ""),
            "live_field_name": str(raw.get("live_field_name") or ""),
            "live_field_label": str(raw.get("live_field_label") or ""),
            "page_id": str(raw.get("page_id") or ""),
            "page_url": str(raw.get("page_url") or ""),
            "message": str(raw.get("message") or ""),
            "screenshot": str(raw.get("screenshot_path") or raw.get("screenshot") or ""),
        })

    qa_summary = result.get("qa_summary") if isinstance(result.get("qa_summary"), dict) else {}
    discrepancy_counts = result.get("discrepancy_counts") if isinstance(result.get("discrepancy_counts"), dict) else {}

    if not qa_summary:
        qa_summary = {
            "missing_fields": int(discrepancy_counts.get("missing_recorded_fields", 0) or 0),
            "new_fields": int(discrepancy_counts.get("new_unexpected_fields", 0) or 0),
            "renamed_fields": int(discrepancy_counts.get("renamed_recorded_fields", 0) or 0),
            "healed_matches": int(discrepancy_counts.get("healed_selector_matches", 0) or 0),
            "not_filled": int(discrepancy_counts.get("recorded_fields_not_filled", 0) or 0),
            "warnings": int(discrepancy_counts.get("warnings", 0) or 0),
            "total_discrepancies": int(discrepancy_counts.get("total", 0) or 0),
        }

    qa_outcome = str(result.get("qa_outcome") or "") or _derive_outcome(result, qa_summary)
    run_metrics = _resolve_run_metrics(result, qa_summary)
    instability_indicators = result.get("instability_indicators") if isinstance(result.get("instability_indicators"), dict) else {}
    if not instability_indicators:
        instability_indicators = _build_instability_indicators(discrepancies)

    trend_ready = result.get("trend_ready") if isinstance(result.get("trend_ready"), dict) else {}
    if not trend_ready:
        trend_ready = {
            "schema": "qa_run_metrics_v1",
            "replay_mode": str(result.get("replay_mode") or qa_summary.get("replay_mode") or "standard"),
            "qa_outcome": qa_outcome,
            "run_metrics": run_metrics,
            "severity_counts": {
                "blocker": int(qa_summary.get("blockers", 0) or 0),
                "warning": int(qa_summary.get("warnings", 0) or 0),
                "info": int(qa_summary.get("info", 0) or 0),
            },
            "discrepancy_counts": {
                "missing_fields": int(qa_summary.get("missing_fields", 0) or 0),
                "new_fields": int(qa_summary.get("new_fields", 0) or 0),
                "renamed_fields": int(qa_summary.get("renamed_fields", 0) or 0),
                "healed_matches": int(qa_summary.get("healed_matches", 0) or 0),
                "not_filled": int(qa_summary.get("not_filled", 0) or 0),
            },
        }

    payload = {
        "workflow_name": workflow_name,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "qa_outcome": qa_outcome,
        "qa_summary": qa_summary,
        "run_metrics": run_metrics,
        "instability_indicators": instability_indicators,
        "trend_ready": trend_ready,
        "discrepancies": discrepancies,
        "playback_status": str(result.get("status") or ""),
        "steps_executed": int(result.get("steps_executed", 0) or 0),
        "steps_failed": int(result.get("steps_failed", 0) or 0),
        "steps_skipped": int(result.get("steps_skipped", 0) or 0),
    }

    with open(paths["json"], "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)

    _write_discrepancy_excel(paths["excel"], payload)

    return {
        "excel": os.path.abspath(paths["excel"]),
        "json": os.path.abspath(paths["json"]),
        "qa_outcome": qa_outcome,
        "qa_summary": qa_summary,
        "run_metrics": run_metrics,
    }


def _resolve_run_metrics(result: Dict[str, Any], qa_summary: Dict[str, Any]) -> Dict[str, Any]:
    metrics = result.get("run_metrics") if isinstance(result.get("run_metrics"), dict) else {}
    if metrics:
        return {
            "pages_compared": int(metrics.get("pages_compared", 0) or 0),
            "fields_scanned": int(metrics.get("fields_scanned", 0) or 0),
            "healed_matches": int(metrics.get("healed_matches", 0) or 0),
            "warnings": int(metrics.get("warnings", 0) or 0),
            "blockers": int(metrics.get("blockers", 0) or 0),
            "runtime_duration_seconds": float(metrics.get("runtime_duration_seconds", 0.0) or 0.0),
        }

    return {
        "pages_compared": int(result.get("pages_compared", 0) or 0),
        "fields_scanned": int(result.get("fields_scanned", 0) or 0),
        "healed_matches": int(qa_summary.get("healed_matches", 0) or 0),
        "warnings": int(qa_summary.get("warnings", 0) or 0),
        "blockers": int(qa_summary.get("blockers", 0) or 0),
        "runtime_duration_seconds": float(result.get("duration", 0.0) or 0.0),
    }


def _build_instability_indicators(discrepancies: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    page_counts: Dict[str, Dict[str, Any]] = {}
    missing_field_counts: Dict[str, Dict[str, Any]] = {}

    for item in discrepancies:
        page_key = str(item.get("page_id") or item.get("page_url") or "unknown_page").strip() or "unknown_page"
        page_entry = page_counts.setdefault(
            page_key,
            {"page_id": str(item.get("page_id") or ""), "page_url": str(item.get("page_url") or ""), "count": 0, "blockers": 0, "warnings": 0},
        )
        page_entry["count"] += 1

        severity = str(item.get("severity") or "warning").strip().lower()
        if severity == "blocker":
            page_entry["blockers"] += 1
        elif severity == "warning":
            page_entry["warnings"] += 1

        if str(item.get("kind") or "") != "missing_recorded_field":
            continue

        field_key = (
            str(item.get("field_id") or "").strip()
            or str(item.get("field_name") or "").strip()
            or str(item.get("field_label") or "").strip()
            or "unknown_field"
        )
        field_entry = missing_field_counts.setdefault(
            field_key,
            {
                "field_key": field_key,
                "field_label": str(item.get("field_label") or ""),
                "field_id": str(item.get("field_id") or ""),
                "field_name": str(item.get("field_name") or ""),
                "count": 0,
            },
        )
        field_entry["count"] += 1

    pages_with_most_discrepancies = sorted(
        page_counts.values(),
        key=lambda row: (int(row.get("count", 0)), int(row.get("blockers", 0))),
        reverse=True,
    )[:5]

    fields_frequently_missing = sorted(
        missing_field_counts.values(),
        key=lambda row: int(row.get("count", 0)),
        reverse=True,
    )[:5]

    return {
        "pages_with_most_discrepancies": pages_with_most_discrepancies,
        "fields_frequently_missing": fields_frequently_missing,
    }


def _severity_from_kind(kind: str) -> str:
    mapping = {
        "missing_recorded_field": "blocker",
        "recorded_field_not_filled": "blocker",
        "new_unexpected_field": "warning",
        "renamed_recorded_field": "warning",
        "healed_selector_match": "info",
    }
    return mapping.get((kind or "").strip().lower(), "warning")


def _derive_outcome(result: Dict[str, Any], qa_summary: Dict[str, Any]) -> str:
    finished = str(result.get("status") or "") == "finished"
    steps_failed = int(result.get("steps_failed", 0) or 0)
    blockers = int(qa_summary.get("blockers", 0) or 0)
    warnings = int(qa_summary.get("warnings", 0) or 0)

    if not finished or steps_failed > 0 or blockers > 0:
        return "Failed"
    if warnings > 0:
        return "Passed with warnings"
    return "Passed"


def _write_discrepancy_excel(path: str, payload: Dict[str, Any]) -> None:
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "QA Summary"

    qa_summary = payload.get("qa_summary") if isinstance(payload.get("qa_summary"), dict) else {}
    run_metrics = payload.get("run_metrics") if isinstance(payload.get("run_metrics"), dict) else {}
    summary_rows = [
        ("Workflow", payload.get("workflow_name", "")),
        ("Generated At", payload.get("generated_at", "")),
        ("QA Outcome", payload.get("qa_outcome", "")),
        ("Replay Mode", qa_summary.get("replay_mode", "")),
        ("Missing Fields", qa_summary.get("missing_fields", 0)),
        ("New Fields", qa_summary.get("new_fields", 0)),
        ("Renamed Fields", qa_summary.get("renamed_fields", 0)),
        ("Healed Matches", qa_summary.get("healed_matches", 0)),
        ("Recorded Not Filled", qa_summary.get("not_filled", 0)),
        ("Warnings", qa_summary.get("warnings", 0)),
        ("Blockers", qa_summary.get("blockers", 0)),
        ("Pages Compared", run_metrics.get("pages_compared", 0)),
        ("Fields Scanned", run_metrics.get("fields_scanned", 0)),
        ("Runtime (seconds)", run_metrics.get("runtime_duration_seconds", 0.0)),
        ("Total Discrepancies", qa_summary.get("total_discrepancies", 0)),
    ]

    for row_idx, (label, value) in enumerate(summary_rows, start=1):
        c1 = ws_summary.cell(row=row_idx, column=1, value=label)
        c2 = ws_summary.cell(row=row_idx, column=2, value=value)
        c1.font = _FONT_META_LBL
        c2.font = _FONT_NORMAL
        c1.fill = _FILL_META
        c2.fill = _FILL_META
        c1.border = _THIN_BORDER
        c2.border = _THIN_BORDER
        c1.alignment = Alignment(horizontal="left")
        c2.alignment = Alignment(horizontal="left")

    ws_summary.column_dimensions["A"].width = 28
    ws_summary.column_dimensions["B"].width = 46

    ws_disc = wb.create_sheet("Discrepancies")
    headers = [
        "Severity",
        "Kind",
        "Field Label",
        "Field ID",
        "Field Name",
        "Tag",
        "Input Type",
        "Required",
        "Visible",
        "Actionable",
        "Expected ID",
        "Expected Name",
        "Live ID",
        "Live Name",
        "Page ID",
        "Page URL",
        "Message",
        "Screenshot",
    ]
    ws_disc.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws_disc.cell(row=1, column=col_idx)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER

    discrepancies = payload.get("discrepancies") if isinstance(payload.get("discrepancies"), list) else []
    for item in discrepancies:
        if not isinstance(item, dict):
            continue
        severity = str(item.get("severity") or "warning").strip().lower()
        fill = _SEVERITY_FILLS.get(severity, _FILL_META)
        ws_disc.append([
            severity,
            item.get("kind", ""),
            item.get("field_label", ""),
            item.get("field_id", ""),
            item.get("field_name", ""),
            item.get("tag", ""),
            item.get("input_type", ""),
            item.get("field_required", False),
            item.get("field_visible", True),
            item.get("field_actionable", True),
            item.get("expected_field_id", ""),
            item.get("expected_field_name", ""),
            item.get("live_field_id", ""),
            item.get("live_field_name", ""),
            item.get("page_id", ""),
            item.get("page_url", ""),
            item.get("message", ""),
            item.get("screenshot", ""),
        ])
        row_num = ws_disc.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws_disc.cell(row=row_num, column=col_idx)
            cell.fill = fill
            cell.border = _THIN_BORDER
            cell.font = _FONT_NORMAL
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [12, 28, 28, 24, 28, 14, 16, 12, 12, 14, 28, 28, 28, 28, 24, 54, 60, 52]
    for col_idx, width in enumerate(widths, start=1):
        ws_disc.column_dimensions[get_column_letter(col_idx)].width = width

    ws_disc.freeze_panes = "A2"

    ws_inst = wb.create_sheet("Instability")
    ws_inst.append(["Pages With Most Discrepancies"])
    ws_inst.cell(row=1, column=1).font = _FONT_HEADER
    ws_inst.cell(row=1, column=1).fill = _FILL_HEADER

    ws_inst.append(["Page ID", "Page URL", "Count", "Blockers", "Warnings"])
    for col_idx in range(1, 6):
        cell = ws_inst.cell(row=2, column=col_idx)
        cell.font = _FONT_META_LBL
        cell.fill = _FILL_META
        cell.border = _THIN_BORDER

    instability = payload.get("instability_indicators") if isinstance(payload.get("instability_indicators"), dict) else {}
    top_pages = instability.get("pages_with_most_discrepancies") if isinstance(instability.get("pages_with_most_discrepancies"), list) else []
    row = 3
    for item in top_pages:
        if not isinstance(item, dict):
            continue
        ws_inst.append([
            item.get("page_id", ""),
            item.get("page_url", ""),
            item.get("count", 0),
            item.get("blockers", 0),
            item.get("warnings", 0),
        ])
        for col_idx in range(1, 6):
            cell = ws_inst.cell(row=row, column=col_idx)
            cell.font = _FONT_NORMAL
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row += 1

    row += 1
    ws_inst.append(["Fields Frequently Missing"])
    ws_inst.cell(row=row, column=1).font = _FONT_HEADER
    ws_inst.cell(row=row, column=1).fill = _FILL_HEADER
    row += 1

    ws_inst.append(["Field Label", "Field ID", "Field Name", "Count"])
    for col_idx in range(1, 5):
        cell = ws_inst.cell(row=row, column=col_idx)
        cell.font = _FONT_META_LBL
        cell.fill = _FILL_META
        cell.border = _THIN_BORDER
    row += 1

    top_missing = instability.get("fields_frequently_missing") if isinstance(instability.get("fields_frequently_missing"), list) else []
    for item in top_missing:
        if not isinstance(item, dict):
            continue
        ws_inst.append([
            item.get("field_label", ""),
            item.get("field_id", ""),
            item.get("field_name", ""),
            item.get("count", 0),
        ])
        for col_idx in range(1, 5):
            cell = ws_inst.cell(row=row, column=col_idx)
            cell.font = _FONT_NORMAL
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row += 1

    ws_inst.column_dimensions["A"].width = 34
    ws_inst.column_dimensions["B"].width = 52
    ws_inst.column_dimensions["C"].width = 30
    ws_inst.column_dimensions["D"].width = 12
    ws_inst.column_dimensions["E"].width = 12

    wb.save(path)


import re  # noqa: E402 — kept at bottom to avoid circular at module level
